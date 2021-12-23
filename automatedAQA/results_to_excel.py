
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import easygui_qt

filter_dic = {
    #Body Coil SNR
    'BC_SNR_TRA':['bc','snr','tra', 'results'],
    'BC_SNR_COR':['bc','snr','cor','results'],
    'BC_SNR_SAG':['bc','snr','sag','results'],
    #Head&Neck COil SNR
    'HNC_SNR_TRA':['hnc','snr','tra','results'],
    'HNC_SNR_COR':['hnc','snr','cor','results'],
    'HNC_SNR_SAG':['hnc','snr','sag','results'],

     #Body Coil Uniformity
    'BC_UNIFORMITY_TRA':['bc','uniformity','tra'],
    'BC_UNIFORMITY_COR':['bc','uniformity','cor'],
    'BC_UNIFORMITY_SAG':['bc','uniformity','sag'],
    #HN Coil Uniformity
    'HNC_UNIFORMITY_TRA':['hnc','uniformity','tra'],
    'HNC_UNIFORMITY_COR':['hnc','uniformity','cor'],
    'HNC_UNIFORMITY_SAG':['hnc','uniformity','sag'],

    # Geometric Linearity
    'GEO_TRA':['geometric_lin','tra'],
    'GEO_COR':['geometric_lin','cor'],
    'GEO_SAG':['geometric_lin','sag'],

    # Slice  width
    'WIDTH_TRA':['slice_width','tra'],
    'WIDTH_COR':['slice_width','cor'],
    'WIDTH_SAG':['slice_width','sag'],

    #Slice Position
    'SlicePosition':['slice_pos'],

    #Ghosting
    'GHOSTING':['ghosting'],
    
    
}
# print(filter_dic)



position_dic = {
    #Body Coil SNR
    'BC_SNR_TRA':'I21',
    'BC_SNR_COR':'A21',
    'BC_SNR_SAG':'E21',
    #Head&Neck COil SNR
    'HNC_SNR_TRA':'I6',
    'HNC_SNR_COR':'A6',
    'HNC_SNR_SAG':'E6',

     #Body Coil Uniformity
    'BC_UNIFORMITY_TRA':'AN6',
    'BC_UNIFORMITY_COR':'AD6',
    'BC_UNIFORMITY_SAG':'AI6',
    #HN Coil Uniformity
    'HNC_UNIFORMITY_TRA':'X6',
    'HNC_UNIFORMITY_COR':'N6',
    'HNC_UNIFORMITY_SAG':'S6',

    # Geometric Linearity
    'GEO_TRA':'BQ6',
    'GEO_COR':'BG6',
    'GEO_SAG':'BL6',

    # Slice  width
    'WIDTH_TRA':'BB6',
    'WIDTH_COR':'AT6',
    'WIDTH_SAG':'AX6',

    #Slice Position
    'SlicePosition':'CB6',

    #Ghosting
    'GHOSTING':'BW6',
    }


def filter_files(filter,files):
    new_files = [file for file in files if filter in str(file.stem).lower()]  
    return new_files    


def find_file(filters,dir):
    files = Path(dir).rglob('*.csv')
    for n,filter in enumerate(filters):
        files = filter_files(filter,files)
        # print(f'\nfilter pass {n}:{files}')
    if len(files) == 1:
        return Path(files[0])  
    elif len(files) > 1:
        [print(file.name) for file in files]
        raise Exception('More than one file matching description.')   
    elif len(files) < 1:
        [print(file.name) for file in files]
        raise Exception('No files found matching description.')   












def cellpos_to_rowcol(position):
    alph = 'abcdefghijklmnopqrstuvwxyz'
    nums = [1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26]
    alph_nums = {a:n for a,n in zip(alph,nums)}
    if position[-2].isnumeric():
        row = int(position[-2:])-1
        code = position.lower()[0:-2]
    else:
        row = int(position[-1])-1
        code = position.lower()[0:-1]

    # print(f'code:{code}')    

    if len(code)==2:
        col = 26*alph_nums[code[0]]+alph_nums[code[1]]-1
    elif len(code)==1:
        col = alph_nums[code[0]]-1
    else:
        raise Exception('Too many rows') 
    # print(col)    
    return (row,col)   



def write_to_excel(df, file, **kwargs):
    try:
        book = load_workbook(file)
        writer = pd.ExcelWriter(file, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df.to_excel(writer, **kwargs)
        writer.save()
    except FileNotFoundError as e:
        df.to_excel(file, **kwargs)


def main(search_in, filter_dic=filter_dic):
    results_sheet = Path(search_in)/'FIJI_Results.xlsx'
    # template_sheet='/Users/papo/Sync/Projects/Annual_QA_pipeline/myexcel.xlsx'

    for key in filter_dic:
        file = find_file(filter_dic[key],search_in)  
        df = pd.read_csv(file)
        position = position_dic[key]
        print(f'Copying the {key} results to position {position} from file:\n.../{file.name}\n')
        row, col = cellpos_to_rowcol(position)
        write_to_excel(df, str(results_sheet), sheet_name="FIJI_Results", index=False, startrow=row, startcol=col)

    # print(f'Copying all FIJI results into the main excel sheet...')
    # fiji = pd.read_excel(results_sheet)
    # write_to_excel(fiji,template_sheet, sheet_name="FIJI_Results", index=False, header=False)
    easygui_qt.show_message(message=f'Done!\n\nNow copy the results in:\n {results_sheet} into the Template excel sheet.')


if __name__ == "__main__":
    main()

# book = load_workbook(results_sheet)
# writer = pd.ExcelWriter(results_sheet, engine='openpyxl') 
# writer.book = book
# writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
# for key in filter_dic:
#     csv_file = find_file(filter_dic[key],search_in)  
#     df = pd.read_csv(csv_file)
#     position = position_dic[key]
#     print(f'Copying the {key} results to position {position} from file:\n.../{csv_file.name}\n')
#     row, col = cellpos_to_rowcol(position)
#     df.to_excel(writer, results_sheet, sheet_name="FIJI_Results", index=False, startrow=row, startcol=col)
# writer.save()


# wb = Workbook()
# ws = wb.active
# ws.title = "FIJI Results"
